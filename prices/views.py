import openpyxl
import requests
from rest_framework import status
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from prices.models import User, FileDetail
from prices.serializers import UserSignUpSerializer, FileProcessSerializer


class UserSignUpAPIView(CreateAPIView):
    """
        This class for register users.
        Users can log in and upload Excel files.
    """
    serializer_class = UserSignUpSerializer
    queryset = User.objects.all()


class FileProcessAPIView(CreateAPIView):
    """
    With this APIview we get Excel file with asset IDs and return the prices of them
    """
    serializer_class = FileProcessSerializer
    permission_classes = (IsAuthenticated,)
    queryset = FileDetail.objects.all()

    def post(self, request, *args, **kwargs):
        serializer = FileProcessSerializer(data=request.data)
        if serializer.is_valid():
            # Here we get current user
            serializer.validated_data['user'] = request.user
            # Here we save the uploaded file and another data's to do something's in continue
            serializer.save()
            # Here we store the uploaded Excel file information in a Python list
            work_book = openpyxl.load_workbook(f"media/uploads/{serializer.validated_data['file']}")
            work_sheet = work_book.active
            # All data uploaded in Excel file is will include in this list
            excel_file_information = []

            # With this loop the data is read from uploaded Excel file and added to excel_file_information
            for i in range(0, work_sheet.max_row):
                for col in work_sheet.iter_cols(1, work_sheet.max_column):
                    excel_file_information.append(col[i].value)

            # This url is an api to get the price of all currencies
            url = 'https://rest.coinapi.io/v1/assets'

            """
            tip:
                Instead of one request and receiving all the prices,
                we could have several requests to the website depending on the names in the Excel file.
                like: https://rest.coinapi.io/v1/exchangerate/BTC/USD or any thing
                But, In my opinion, by doing this, we will make several requests to the website and our
                work speed will be reduced
            """

            # Here we put the API key in the request header to the url
            headers = {'X-CoinAPI-Key': '45A8EC07-C27C-4C00-9B77-B42EEED37541'}
            currency_prices = requests.get(url, headers=headers)
            currency_prices = currency_prices.json()
            result = {}

            # And here process the data
            for i in currency_prices:
                if i['asset_id'] in excel_file_information or i['name'] in excel_file_information:
                    if not 'price_usd' in i:
                        result[i['asset_id']] = f"Currency price {i['asset_id']} not found"
                    else:
                        result[i['asset_id']] = i['price_usd']

            return Response({"result": result, "data": serializer.data, "status": status.HTTP_201_CREATED,
                             "errors": serializer.errors})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
